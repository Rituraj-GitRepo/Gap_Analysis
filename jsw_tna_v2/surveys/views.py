import io
import json
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404

from .models import FunctionalHeadResponse, SelfAssessmentResponse
from .competencies import FUNCTIONS, SURVEY_META, PROFICIENCY_LEVELS, IMPORTANCE_LEVELS


# ── helpers ───────────────────────────────────────────────────────────────────

def get_client_ip(request):
    x = request.META.get('HTTP_X_FORWARDED_FOR')
    return x.split(',')[0].strip() if x else request.META.get('REMOTE_ADDR')


def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.session.get('admin_logged_in'):
            return redirect('/admin-login/')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


def group_competencies(competencies):
    """Split list into tech / behavioural groups for templates."""
    tech = [c for c in competencies if c['category'] == 'Technical']
    beh  = [c for c in competencies if c['category'] == 'Behavioral']
    return tech, beh


# ── public views ──────────────────────────────────────────────────────────────

def welcome(request):
    if request.method == 'POST':
        name        = request.POST.get('name', '').strip()
        emp_code    = request.POST.get('emp_code', '').strip()
        role        = request.POST.get('role', '').strip()
        function_key= request.POST.get('function', '').strip()

        errors = []
        if not name:        errors.append('Employee name is required.')
        if not emp_code:    errors.append('Employee code is required.')
        if not role:        errors.append('Please select your role.')
        if not function_key:errors.append('Please select your function.')
        if function_key and function_key not in FUNCTIONS:
            errors.append('Invalid function selected.')

        if errors:
            return render(request, 'surveys/welcome.html', {
                'error': ' '.join(errors),
                'functions': FUNCTIONS,
                'name': name, 'emp_code': emp_code,
                'role': role, 'function_key': function_key,
            })

        request.session['user_name']     = name
        request.session['user_emp_code'] = emp_code
        request.session['user_role']     = role
        request.session['user_function'] = function_key

        if role == 'FH':
            return redirect(f'/survey/functional-head/{function_key}/')
        return redirect(f'/survey/self-assessment/{function_key}/')

    return render(request, 'surveys/welcome.html', {'functions': FUNCTIONS})


def functional_head_survey(request, function_key):
    if function_key not in FUNCTIONS:
        return redirect('/')

    func_data = FUNCTIONS[function_key]
    tech_comps, beh_comps = group_competencies(func_data['competencies'])
    prefill = {
        'name':     request.session.get('user_name', ''),
        'emp_code': request.session.get('user_emp_code', ''),
        'role':     request.session.get('user_role', 'FH'),
    }

    if request.method == 'POST':
        try:
            ratings = {}
            for comp in func_data['competencies']:
                code       = comp['code']
                importance = request.POST.get(f'importance_{code}', '').strip()
                top10      = request.POST.get(f'top10_{code}') == 'on'
                day1_level = request.POST.get(f'day1_{code}', '').strip()
                if importance:
                    ratings[code] = {
                        'name':      comp['name'],
                        'category':  comp['category'],
                        'kpi_link':  comp['kpi_link'],
                        'importance': int(importance),
                        'top10':     top10,
                        'day1_level': int(day1_level) if day1_level else None,
                    }

            obj = FunctionalHeadResponse(
                employee_name=request.POST.get('employee_name', '').strip(),
                employee_code=request.POST.get('employee_code', '').strip(),
                role='FH',
                function=function_key,
                date=request.POST.get('date'),
                competency_ratings=ratings,
                q1_critical_competency=request.POST.get('q1', '').strip(),
                q2_gap_risk=request.POST.get('q2', '').strip(),
                q3_alp_theme=request.POST.get('q3', '').strip(),
                ip_address=get_client_ip(request),
            )
            obj.save()
            return redirect(f'/survey/success/?type=functional-head&function={function_key}')
        except Exception as e:
            return render(request, 'surveys/functional_head_survey.html', {
                'function_key': function_key, 'func_data': func_data,
                'tech_comps': tech_comps, 'beh_comps': beh_comps,
                'proficiency_levels': PROFICIENCY_LEVELS,
                'importance_levels': IMPORTANCE_LEVELS,
                'error': f'Submission failed: {e}', 'prefill': prefill,
            })

    return render(request, 'surveys/functional_head_survey.html', {
        'function_key': function_key, 'func_data': func_data,
        'tech_comps': tech_comps, 'beh_comps': beh_comps,
        'proficiency_levels': PROFICIENCY_LEVELS,
        'importance_levels': IMPORTANCE_LEVELS,
        'prefill': prefill,
    })


def self_assessment(request, function_key):
    if function_key not in FUNCTIONS:
        return redirect('/')

    func_data = FUNCTIONS[function_key]
    tech_comps, beh_comps = group_competencies(func_data['competencies'])
    prefill = {
        'name':     request.session.get('user_name', ''),
        'emp_code': request.session.get('user_emp_code', ''),
        'role':     request.session.get('user_role', ''),
    }

    if request.method == 'POST':
        try:
            ratings = {}
            for comp in func_data['competencies']:
                code          = comp['code']
                current_level = request.POST.get(f'level_{code}', '').strip()
                evidence      = request.POST.get(f'evidence_{code}', '').strip()
                if current_level:
                    ratings[code] = {
                        'name':         comp['name'],
                        'category':     comp['category'],
                        'kpi_link':     comp['kpi_link'],
                        'current_level': int(current_level),
                        'evidence':     evidence,
                    }

            obj = SelfAssessmentResponse(
                employee_name=request.POST.get('employee_name', '').strip(),
                employee_code=request.POST.get('employee_code', '').strip(),
                role=request.POST.get('role', '').strip(),
                function=function_key,
                date=request.POST.get('date'),
                competency_ratings=ratings,
                q1_strengths=request.POST.get('q1', '').strip(),
                q2_development_areas=request.POST.get('q2', '').strip(),
                ip_address=get_client_ip(request),
            )
            obj.save()
            return redirect(f'/survey/success/?type=self-assessment&function={function_key}')
        except Exception as e:
            return render(request, 'surveys/self_assessment.html', {
                'function_key': function_key, 'func_data': func_data,
                'tech_comps': tech_comps, 'beh_comps': beh_comps,
                'proficiency_levels': PROFICIENCY_LEVELS,
                'error': f'Submission failed: {e}', 'prefill': prefill,
            })

    return render(request, 'surveys/self_assessment.html', {
        'function_key': function_key, 'func_data': func_data,
        'tech_comps': tech_comps, 'beh_comps': beh_comps,
        'proficiency_levels': PROFICIENCY_LEVELS,
        'prefill': prefill,
    })


def success(request):
    survey_type = request.GET.get('type', 'survey')
    function_key = request.GET.get('function', '')
    meta = SURVEY_META.get(survey_type, {'title': 'Survey', 'icon': '✅', 'color': '#16a34a', 'subtitle': ''})
    return render(request, 'surveys/success.html', {
        'meta': meta,
        'survey_type': survey_type,
        'func_label': FUNCTIONS.get(function_key, {}).get('label', ''),
        'user_name': request.session.get('user_name', ''),
    })


# ── admin auth ────────────────────────────────────────────────────────────────

def admin_login(request):
    if request.session.get('admin_logged_in'):
        return redirect('/admin-dashboard/')
    error = None
    if request.method == 'POST':
        if (request.POST.get('username', '').strip() == settings.ADMIN_USERNAME and
                request.POST.get('password', '').strip() == settings.ADMIN_PASSWORD):
            request.session['admin_logged_in'] = True
            return redirect('/admin-dashboard/')
        error = 'Invalid credentials. Please try again.'
    return render(request, 'surveys/admin_login.html', {'error': error})


def admin_logout(request):
    request.session.flush()
    return redirect('/admin-login/')


# ── admin dashboard ───────────────────────────────────────────────────────────

@admin_required
def admin_dashboard(request):
    fh_total = FunctionalHeadResponse.objects.count()
    sa_total = SelfAssessmentResponse.objects.count()

    func_stats = []
    for fk, fd in FUNCTIONS.items():
        fh_c = FunctionalHeadResponse.objects.filter(function=fk).count()
        sa_c = SelfAssessmentResponse.objects.filter(function=fk).count()
        func_stats.append({'key': fk, 'label': fd['label'],
                           'fh_count': fh_c, 'sa_count': sa_c, 'total': fh_c + sa_c})

    return render(request, 'surveys/admin_dashboard.html', {
        'fh_total': fh_total, 'sa_total': sa_total,
        'total': fh_total + sa_total,
        'get_sa': SelfAssessmentResponse.objects.filter(role='GET').count(),
        'mt_sa':  SelfAssessmentResponse.objects.filter(role='MT').count(),
        'func_stats': func_stats,
        'recent_fh': FunctionalHeadResponse.objects.order_by('-submitted_at')[:5],
        'recent_sa': SelfAssessmentResponse.objects.order_by('-submitted_at')[:5],
        'functions': FUNCTIONS,
    })


# ── admin responses ───────────────────────────────────────────────────────────

@admin_required
def admin_responses(request, survey_type):
    function_filter = request.GET.get('function', '')
    role_filter     = request.GET.get('role', '')

    if survey_type == 'functional-head':
        qs = FunctionalHeadResponse.objects.all()
        if function_filter: qs = qs.filter(function=function_filter)
        title = 'Functional Head Survey Responses'
    elif survey_type == 'self-assessment':
        qs = SelfAssessmentResponse.objects.all()
        if function_filter: qs = qs.filter(function=function_filter)
        if role_filter in ('GET', 'MT'): qs = qs.filter(role=role_filter)
        title = 'Employee Self-Assessment Responses'
    else:
        return redirect('/admin-dashboard/')

    return render(request, 'surveys/admin_responses.html', {
        'responses': qs, 'title': title,
        'survey_type': survey_type,
        'function_filter': function_filter, 'role_filter': role_filter,
        'functions': FUNCTIONS,
    })


# ── delete response ───────────────────────────────────────────────────────────

@admin_required
def delete_response(request, survey_type, pk):
    if request.method == 'POST':
        if survey_type == 'functional-head':
            get_object_or_404(FunctionalHeadResponse, pk=pk).delete()
        elif survey_type == 'self-assessment':
            get_object_or_404(SelfAssessmentResponse, pk=pk).delete()
    return redirect(f'/admin-responses/{survey_type}/')


# ── gap analysis ──────────────────────────────────────────────────────────────

@admin_required
def gap_analysis(request):
    function_key = request.GET.get('function', 'finance')
    if function_key not in FUNCTIONS:
        function_key = list(FUNCTIONS.keys())[0]

    func_data = FUNCTIONS[function_key]
    comps = func_data['competencies']
    comp_codes = [c['code'] for c in comps]
    comp_names = [c['name'] for c in comps]

    # ── FH averages: importance and day-1 level ───────────────────────────────
    fh_qs = FunctionalHeadResponse.objects.filter(function=function_key)
    fh_importance  = defaultdict(list)
    fh_day1        = defaultdict(list)
    fh_top10_count = defaultdict(int)
    fh_total = fh_qs.count()

    for resp in fh_qs:
        for code, data in resp.competency_ratings.items():
            if data.get('importance'):
                fh_importance[code].append(data['importance'])
            if data.get('day1_level'):
                fh_day1[code].append(data['day1_level'])
            if data.get('top10'):
                fh_top10_count[code] += 1

    # ── SA averages: current proficiency ─────────────────────────────────────
    sa_qs = SelfAssessmentResponse.objects.filter(function=function_key)
    sa_levels = defaultdict(list)
    sa_total  = sa_qs.count()

    for resp in sa_qs:
        for code, data in resp.competency_ratings.items():
            if data.get('current_level'):
                sa_levels[code].append(data['current_level'])

    # ── Build chart-ready data ────────────────────────────────────────────────
    chart_data = []
    for comp in comps:
        code = comp['code']
        avg_imp   = round(sum(fh_importance[code]) / len(fh_importance[code]), 2) if fh_importance[code] else 0
        avg_day1  = round(sum(fh_day1[code]) / len(fh_day1[code]), 2) if fh_day1[code] else 0
        avg_sa    = round(sum(sa_levels[code]) / len(sa_levels[code]), 2) if sa_levels[code] else 0
        gap       = round(avg_day1 - avg_sa, 2)
        top10_pct = round((fh_top10_count[code] / fh_total * 100), 1) if fh_total else 0
        chart_data.append({
            'code': code, 'name': comp['name'], 'category': comp['category'],
            'kpi_link': comp['kpi_link'],
            'avg_importance': avg_imp,
            'avg_day1': avg_day1,
            'avg_sa': avg_sa,
            'gap': gap,
            'top10_pct': top10_pct,
            'top10_count': fh_top10_count[code],
            'gap_severity': 'high' if gap >= 2 else ('medium' if gap >= 1 else 'low'),
        })

    chart_data_json = json.dumps(chart_data)

    # Summary stats
    high_gaps  = [d for d in chart_data if d['gap_severity'] == 'high']
    top10_list = sorted(chart_data, key=lambda x: x['top10_count'], reverse=True)[:10]

    return render(request, 'surveys/gap_analysis.html', {
        'function_key':    function_key,
        'func_data':       func_data,
        'functions':       FUNCTIONS,
        'chart_data':      chart_data,
        'chart_data_json': chart_data_json,
        'fh_total':        fh_total,
        'sa_total':        sa_total,
        'high_gaps':       high_gaps,
        'top10_list':      top10_list,
        'comp_names_json': json.dumps(comp_names),
        'comp_codes_json': json.dumps(comp_codes),
    })


# ── excel export ──────────────────────────────────────────────────────────────

@admin_required
def export_excel(request, survey_type):
    function_filter = request.GET.get('function', '')
    role_filter     = request.GET.get('role', '')

    wb = openpyxl.Workbook()

    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', fgColor='1A2D5A')
    alt_fill  = PatternFill('solid', fgColor='E8F0FE')
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin      = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'),  bottom=Side(style='thin'))

    def sh(ws, row, cols):
        for col in range(1, cols + 1):
            c = ws.cell(row=row, column=col)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = center; c.border = thin

    def sr(ws, row, cols, alt=False):
        for col in range(1, cols + 1):
            c = ws.cell(row=row, column=col)
            if alt: c.fill = alt_fill
            c.alignment = left_wrap; c.border = thin

    if survey_type == 'functional-head':
        qs = FunctionalHeadResponse.objects.all()
        if function_filter: qs = qs.filter(function=function_filter)

        ws = wb.active; ws.title = 'FH Responses'
        hdrs = ['#','Name','Emp Code','Role','Function','Date','Top-10 Count',
                'Avg Importance','Q1 Critical Comp','Q2 Gap Risk','Q3 ALP Theme','Submitted At']
        ws.append(hdrs); sh(ws, 1, len(hdrs))
        for i, r in enumerate(qs, 1):
            ws.append([i, r.employee_name, r.employee_code, r.get_role_display(),
                       r.get_function_display(), str(r.date), r.top10_count(), r.avg_importance(),
                       r.q1_critical_competency, r.q2_gap_risk, r.q3_alp_theme,
                       r.submitted_at.strftime('%Y-%m-%d %H:%M')])
            sr(ws, i+1, len(hdrs), alt=(i % 2 == 0))
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 22

        ws2 = wb.create_sheet('FH Competency Detail')
        hdrs2 = ['Name','Emp Code','Function','Code','Competency','Category',
                 'KPI Link','Importance (1-5)','Top 10?','Day-1 Level']
        ws2.append(hdrs2); sh(ws2, 1, len(hdrs2))
        r2 = 2
        for r in qs:
            for code, data in r.competency_ratings.items():
                ws2.append([r.employee_name, r.employee_code, r.get_function_display(),
                             code, data.get('name',''), data.get('category',''),
                             data.get('kpi_link',''), data.get('importance',''),
                             'Yes' if data.get('top10') else 'No', data.get('day1_level','')])
                sr(ws2, r2, len(hdrs2), alt=(r2 % 2 == 0)); r2 += 1
        ws2.column_dimensions['E'].width = 35; ws2.column_dimensions['G'].width = 45

    elif survey_type == 'self-assessment':
        qs = SelfAssessmentResponse.objects.all()
        if function_filter: qs = qs.filter(function=function_filter)
        if role_filter in ('GET','MT'): qs = qs.filter(role=role_filter)

        ws = wb.active; ws.title = 'SA Responses'
        hdrs = ['#','Name','Emp Code','Role','Function','Date',
                'Avg Proficiency','Q1 Strengths','Q2 Dev Areas','Submitted At']
        ws.append(hdrs); sh(ws, 1, len(hdrs))
        for i, r in enumerate(qs, 1):
            ws.append([i, r.employee_name, r.employee_code, r.get_role_display(),
                       r.get_function_display(), str(r.date), r.avg_proficiency(),
                       r.q1_strengths, r.q2_development_areas,
                       r.submitted_at.strftime('%Y-%m-%d %H:%M')])
            sr(ws, i+1, len(hdrs), alt=(i % 2 == 0))
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 22

        ws2 = wb.create_sheet('SA Competency Detail')
        hdrs2 = ['Name','Emp Code','Role','Function','Code','Competency',
                 'Category','KPI Link','Current Level (1-5)','Evidence']
        ws2.append(hdrs2); sh(ws2, 1, len(hdrs2))
        r2 = 2
        for r in qs:
            for code, data in r.competency_ratings.items():
                ws2.append([r.employee_name, r.employee_code, r.get_role_display(),
                             r.get_function_display(), code, data.get('name',''),
                             data.get('category',''), data.get('kpi_link',''),
                             data.get('current_level',''), data.get('evidence','')])
                sr(ws2, r2, len(hdrs2), alt=(r2 % 2 == 0)); r2 += 1
        ws2.column_dimensions['F'].width = 35; ws2.column_dimensions['H'].width = 45

    info = wb.create_sheet('Export Info', 0)
    info['A1'] = 'JSW Motors — Competency & TNA Survey Data Export'
    info['A1'].font = Font(bold=True, size=14, color='1A2D5A')
    info['A2'] = f'Survey Type: {survey_type.replace("-"," ").title()}'
    info['A3'] = f'Function: {FUNCTIONS.get(function_filter,{}).get("label","All")}'
    info['A4'] = f'Role Filter: {role_filter or "All"}'
    info['A5'] = f'Exported At: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    info.column_dimensions['A'].width = 60

    for sheet in wb.worksheets:
        sheet.freeze_panes = sheet.freeze_panes or 'A2'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)

    fl = FUNCTIONS.get(function_filter,{}).get('label','ALL').replace(' & ','_').replace(' ','_')
    fn = f"JSW_TNA_{survey_type}_{fl}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fn}"'
    return resp
